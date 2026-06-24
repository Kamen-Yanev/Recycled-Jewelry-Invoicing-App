from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import date
import os
import tempfile
import subprocess
import shutil
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaFileUpload
import platform

# add warning message when wrong file type

metal_colors_dict = {'AU': PatternFill(start_color = 'd3af37', end_color = 'd3af37', fill_type = 'solid'),
               'AG': PatternFill(start_color = 'e6e6e6', end_color = 'e6e6e6', fill_type = 'solid'),
                'PD': PatternFill(start_color = '989da1', end_color = '989da1', fill_type = 'solid'),
                 'PT': PatternFill(start_color = 'e5e4e2', end_color = 'e5e4e2', fill_type = 'solid')}

def customers(): # returns a dict of client_metal : discount rate (float) Note: 'Rate of day' == 0 so that everything is numeric

    df = pd.read_excel('Updated Balance (1).xlsx', sheet_name = 'Sheet3', dtype = str)
    df = df.apply(lambda x: x.str.strip() if x.dtype == 'object' or x.dtype == 'string' else x)
    df.columns = ['Client', 'AU', 'AG', 'PT', 'PD']

    clients = {}
    for client in range(len(df)):
        clients[f'{df.iloc[client,0]}'] = {} # fill clients dict with client : nested dict pairs
        for metal in range(1,5): # populate nested dictionaries with metal : discount pairs
            if df.iloc[client,metal] == 'Rate of day':
                clients[f'{df.iloc[client,0]}'][f'{df.columns[metal]}'] = 0
            elif pd.isna(df.iloc[client,metal]):
                continue
            else:
                clients[f'{df.iloc[client,0]}'][f'{df.columns[metal]}'] = float(df.iloc[client,metal]) 
    return clients

def get_drive_service():
    scopes = ['https://www.googleapis.com/auth/drive']
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', scopes)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)


def find_or_create_folder(service, name, parent_id):
    # First check if folder already exists
    query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    folders = results.get('files', [])
    for folder in folders:
        if name.lower() in folder['name'].lower():
            return folder['id']

    # Doesn't exist — create it
    metadata = {
        'name': name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder['id']


def upload_to_drive(pdf_path, pn):
    service = get_drive_service()

    # Find Recycled Jewellery root folder
    query = "name='Recycled Jewellery' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id)").execute()
    folders = results.get('files', [])
    if not folders:
        raise Exception("Could not find 'Recycled Jewellery' folder in Drive")
    root_id = folders[0]['id']

    # Navigate / create path: year -> Purchase Invoices -> Month -> CPM -> PN folder
    current_year = date.today().strftime('%Y')
    month_name = date.today().strftime('%B')
    pn_folder_name = f'PN{pn}'

    year_id = find_or_create_folder(service, current_year, root_id)
    purchase_id = find_or_create_folder(service, 'Purchase Invoices', year_id)
    month_id = find_or_create_folder(service, month_name, purchase_id)
    cpm_id = find_or_create_folder(service, 'CPM', month_id)
    pn_id = find_or_create_folder(service, pn_folder_name, cpm_id)

    # Upload the PDF into the PN folder
    file_metadata = {
        'name': f'{pn_folder_name}.pdf',
        'parents': [pn_id]
    }
    media = MediaFileUpload(pdf_path, mimetype='application/pdf')
    service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f"Uploaded {pn_folder_name}.pdf to Drive successfully.")

class App:
    def __init__(self):
        self.root = Tk()
        self.cpm_document = ''
        self.client = '' 
        self.client_rates = customers()
        self.build_ui()
        self.root.mainloop()

    def upload_cpm_doc(self): # holds the path to the cpm document, used to call invoice() (Upload CPM doc button is bound to this)
        # Program asks user to give desired file's path
        path = filedialog.askopenfilename(parent=self.root) 
        if path:
            self.cpm_document = path

    def isolate_cpm_table(self, og_df): # return a list of dicts 
        necessary_info = []

        # Find NO: and SUB TOTAL, use them to build smaller data frame (they are the edges)
        locate_start = np.where(og_df.values == 'NO:')
        start_row = int(locate_start[0][0])

        locate_end = np.where(og_df.values == 'SUB TOTAL')
        end_row = int(locate_end[0][0])

        # build smaller data frame
        df = og_df[start_row:end_row-1]
        df = df.iloc[:, 1:]

        # fix column names 
        df.columns = df.iloc[0] # sets row 1 as column names
        df = df.iloc[1:]
        df = df[df["NO:"].notna()] # get only rows with numbers!
        df = df.drop(columns = df.columns[[6,7,9]]) # drop Refining, quantity due, total due 
        df.columns = ['Number', 'Description', 'Melted Weight', 'Metal', 'ASSAY %', 'Fine Metal',
                    'Rate'] # change column names to match invoice column names
        df = df.astype({'Fine Metal': float, 'Rate': float})

        for i in range(len(df)):
            necessary_info.append(df.iloc[i].to_dict())
        return necessary_info

    def isolate_pn(self, og_df): # returns the pn as str
        pn_position = np.where(og_df.values == 'PN')
        pn_row = int(pn_position[0][0])
        pn_col = int(pn_position[1][0])
        pn = og_df.iloc[pn_row,pn_col+1]
        return pn

    def invoice(self, cpm_doc, client): # main function, handles writing to the invoice, calculations and letting user save file as sth

        og_df = pd.read_excel(cpm_doc, dtype = str) # read in the CPM doc as a data frame
        og_df = og_df.apply(lambda x: x.str.strip() if x.dtype == 'string' or x.dtype == 'object' else x)

        cpm_table = self.isolate_cpm_table(og_df) # list of dictionaries, each dict is the full row of the cpm doc
        pn = self.isolate_pn(og_df) # add to invoice at the end, not in loop

        inv_df = pd.read_excel('Invoice Templates copy.xlsx', sheet_name = 'CPM purchase Invoice', dtype = str)
        inv_df = inv_df.apply(lambda x: x.apply(lambda v: ' '.join(str(v).split()) if isinstance(v, str) else v) if x.dtype == 'object' or x.dtype == 'string' else x)
        starting_row = int(np.where(inv_df.values == 'Melted Weight')[0][0] + 3) # + 3 bc pandas 0-index, excel 1-index
        
        wb = load_workbook('Invoice Templates copy.xlsx', data_only = True)
        ws = wb['CPM purchase Invoice']
        column_names = 'BCDEFG' # Number, Description, Melted Weight, Metal, Assay, Fine Metal
        table_cols = starting_row - 1

        # add pn
        ws['H12'] = f'PN{pn}'

        # update date of invoice to current date
        ws['H11'] = date.today()

        # Write Number:Fine Metal to invoice
        for dict in cpm_table:
            for col in column_names:
                cell_name = ws[f'{col}{table_cols}'].value
                corrected_cell_name = ' '.join(str(cell_name).split()) # some names have whitespaces in between words
                ws[f'{col}{starting_row}'] = dict[corrected_cell_name] # this cell in the invoice becomes the dict value of the key with the same name as the cell
            starting_row += 1
        
        # Determine all unique metals and respective rates in the CPM doc
        cpm_metals = {} # unique rates stored in a dict
        for dict in cpm_table:
            if dict['Rate'] in cpm_metals:
                continue
            else:
                cpm_metals[dict['Metal']] = float(dict['Rate'])

        # Write market rates to invoice
        invoice_market_rates = {'AU': 'C15', 'AG': 'C16', 'AG2': 'C17', 'PT': 'C18', 'PD': 'C19'} # mapping of market rate entry cells in invoice
        invoice_discount_fields = {'AU': 'D15', 'AG': 'D16', 'AG2': 'D17', 'PT': 'D18', 'PD': 'D19'} # mapping of discount entry cells in invoice
        client_information_fields = {'AU': 'I17', 'AG': 'I18', 'AG2': 'I19', 'PT': 'I20', 'PD': 'I21'} # mapping of client information table cells in invoice
        main_table_rate_fields = ['H33', 'H32', 'H31', 'H30', 'H29', 'H28', 'H27', 'H26', 'H25', 'H24'] # mapping of main table rate fields in invoice
        main_table_totals = ['I33', 'I32', 'I31', 'I30', 'I29', 'I28', 'I27', 'I26', 'I25', 'I24'] # mapping of main table Total fields in invoice
        main_table_assay = ['F33', 'F32', 'F31', 'F30', 'F29', 'F28', 'F27', 'F26', 'F25', 'F24']
        main_table_metal = ['E33', 'E32', 'E31', 'E30', 'E29', 'E28', 'E27', 'E26', 'E25', 'E24']

        for metal in cpm_metals:
            if metal in invoice_market_rates:
                ws[invoice_market_rates[metal]] = cpm_metals[metal] # write the market price of the given metal to the invoice 

        # Write client discount % and client adjusted rate to invoice        
        for metal in cpm_metals:
            if metal in self.client_rates[client]:
                if metal in invoice_discount_fields:
                    if self.client_rates[client][metal] == 0:
                        ws[invoice_discount_fields[metal]] = 'Rate of the day' # add rate of day into discount table
                        ws[client_information_fields[metal]] = f'R{cpm_metals[metal]}' # add rate in Client Information table
                    else:
                        ws[invoice_discount_fields[metal]] = f'{self.client_rates[client][metal]*100:.2f}%' # write the client discount percentage to the invoice
                        ws[client_information_fields[metal]] = f'R{cpm_metals[metal]*self.client_rates[client][metal] + cpm_metals[metal]:,.2f}' # add rate in Client Information table
        
        # Write the client rate, total and grand total to the main CPM table in the invoice (have to loop through each table row now)
        grand_total = 0
        assay_cell_pop = '' # to store popped value, need to convert assay cell into percentage
        metal_cell_pop = ''
        for dict in cpm_table:
            ws[main_table_rate_fields.pop()] = f'R{cpm_metals[dict['Metal']]*self.client_rates[client][dict['Metal']] + cpm_metals[dict['Metal']]:,.2f}'
            ws[main_table_totals.pop()] = f'R{dict['Fine Metal'] * (cpm_metals[dict['Metal']]*self.client_rates[client][dict['Metal']] + cpm_metals[dict['Metal']]):,.2f}'
            grand_total += dict['Fine Metal'] * (cpm_metals[dict['Metal']]*self.client_rates[client][dict['Metal']] + cpm_metals[dict['Metal']])

            # show assay up to and including second decimal point
            assay_cell_pop = main_table_assay.pop()
            assay_value = round(float(ws[assay_cell_pop].value) * 100,2)
            ws[assay_cell_pop] = f'{assay_value:.2f}%' 

            # color metal background
            metal_cell_pop = main_table_metal.pop()
            metal_value = ws[metal_cell_pop].value
            if metal_value in metal_colors_dict:
                ws[metal_cell_pop].fill = metal_colors_dict[metal_value]

        # write grand total to invoice
        ws['I34'] = f'R{grand_total:,.2f}' 

        updated_invoice = wb

        # Saving the file
        # Save openpyxl wb to a hidden temporary file 
        # Remove all sheets except CPM purchase Invoice
        sheets_to_delete = [s for s in wb.sheetnames if s != 'CPM purchase Invoice']
        for sheet in sheets_to_delete:
            del wb[sheet]

        from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
        ws.sheet_properties = WorksheetProperties()
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToWidth = 1
        ws.print_area = 'B1:I36'

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        pdf_path = filedialog.asksaveasfilename(
            parent=self.root,
            defaultextension='.pdf',
            filetypes=[('PDF files', '*.pdf')],
            title='Save invoice as PDF',
            initialfile=f'PN{pn}'
        )

        # Ask where to save the Excel file too
        xlsx_path = os.path.splitext(pdf_path)[0] + '.xlsx'

        if pdf_path:
            if platform.system() == 'Windows':
                libreoffice_path = r'C:\Program Files\LibreOffice\program\soffice.exe'
            else:
                libreoffice_path = '/Applications/LibreOffice.app/Contents/MacOS/soffice'

            subprocess.run([
                libreoffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(pdf_path),
                tmp_path
            ])
            converted = os.path.join(os.path.dirname(pdf_path), os.path.basename(tmp_path).replace('.xlsx', '.pdf'))
            os.rename(converted, pdf_path)
            try:                  
                upload_to_drive(pdf_path, pn)
            except Exception as e:
                print(f"Drive upload failed: {e}")
        if xlsx_path:
            shutil.copy(tmp_path, xlsx_path)


        os.remove(tmp_path)
    
    def build_ui(self):
        self.root.geometry('640x200')
        self.root.title('Invoice Helper')

        # Mainframe + upload cpm doc buton
        title_frame = ttk.Frame(self.root)
        title_frame.pack(pady=(20, 0))

        title_label = ttk.Label(title_frame, text='Invoice Helper', font=('TkDefaultFont', 16, 'bold'))
        title_label.pack()

        subtitle_label = ttk.Label(title_frame, text='Generate client invoices from CPM documents', foreground='gray')
        subtitle_label.pack()

        # Separator
        sep_frame = ttk.Frame(self.root)
        sep_frame.pack(fill='x', padx=20, pady=10)
        ttk.Separator(sep_frame, orient='horizontal').pack(fill='x')

        # Controls in their own frame
        mainframe = ttk.Frame(self.root, padding=(20, 0, 20, 20))
        mainframe.pack()

        ttk.Label(mainframe, text='Client:').grid(column=0, row=0, sticky='e', padx=(0, 5))

        # Create combobox widget
        client_rates_list = [i for i in self.client_rates.keys()]
        client_rates_list.append('New client') # add new client_option
        n = StringVar()
        self.combobox = ttk.Combobox(mainframe, width=25, textvariable=n, values=client_rates_list, state='readonly')
        self.combobox.grid(column=1, row=0, padx=(0, 10))

        button_frame = ttk.Frame(mainframe)
        button_frame.grid(column=2, row=0)
        ttk.Button(button_frame, text='Upload CPM', command=self.upload_cpm_doc).grid(column=0, row=0, padx=(0, 5))
        ttk.Button(button_frame, text='Generate Invoice', command=lambda: self.invoice(self.cpm_document, self.client)).grid(column=1, row=0)

        self.combobox.bind('<<ComboboxSelected>>', self.select_client) # Create binding function for combobox

    def select_client(self, event=None):
        if self.combobox.get() == 'New client':
            self.open_new_client_dialog()
        else:
            self.client = self.combobox.get()
    # Function that is run when an entry in combobox is selected - assigns a value to global client (used to call invoice())
    def open_new_client_dialog(self):
    # What happens if you select new client
        top = Toplevel()
        top.grab_set()
        top.geometry("340x270")
        top.title("New client details")
        top.configure(bg="#f9f9f9")
        top.resizable(False, False)

        # Padding frame
        frame = Frame(top, bg="#f9f9f9", padx=18, pady=16)
        frame.pack(fill='both', expand=True)

        # Title + divider
        title_label = Label(frame, text="New client details",
                            font=('TkDefaultFont', 11, 'bold'),
                            bg="#f9f9f9", fg="#1a1a1a")
        title_label.grid(column=0, row=0, columnspan=2, sticky='w', pady=(0, 4))

        separator = Frame(frame, height=1, bg="#e0e0e0")
        separator.grid(column=0, row=1, columnspan=2, sticky='ew', pady=(0, 12))

        # Field rows
        metal_label = Label(frame, text="Metal:", font=('TkDefaultFont', 10),
                            bg="#f9f9f9", fg="#555555", anchor='e', width=9)
        discount_label = Label(frame, text="Discount:", font=('TkDefaultFont', 10),
                            bg="#f9f9f9", fg="#555555", anchor='e', width=9)

        metal_var = StringVar()
        discount_var = StringVar()

        metal_entry = Entry(frame, textvariable=metal_var,
                            relief='flat', bd=1, highlightthickness=1,
                            highlightbackground="#cccccc", highlightcolor="#888888",
                            font=('TkDefaultFont', 10))
        discount_entry = Entry(frame, textvariable=discount_var,
                            relief='flat', bd=1, highlightthickness=1,
                            highlightbackground="#cccccc", highlightcolor="#888888",
                            font=('TkDefaultFont', 10))

        metal_label.grid(column=0, row=2, sticky='e', pady=4)
        metal_entry.grid(column=1, row=2, sticky='ew', padx=(8, 0), pady=4)
        discount_label.grid(column=0, row=3, sticky='e', pady=4)
        discount_entry.grid(column=1, row=3, sticky='ew', padx=(8, 0), pady=4)

        frame.columnconfigure(1, weight=1)

        # Warning box
        warn_frame = Frame(frame, bg="#fffbf0",
                        highlightthickness=1, highlightbackground="#f0d080")
        warn_frame.grid(column=0, row=4, columnspan=2, sticky='ew', pady=(12, 0))

        warn_inner = Frame(warn_frame, bg="#fffbf0", padx=10, pady=8)
        warn_inner.pack(fill='both')

        Label(warn_inner, text="If rate of day, enter 0.",
            font=('TkDefaultFont', 9, 'bold'), bg="#fffbf0", fg="#7a5500",
            anchor='w').pack(fill='x')
        Label(warn_inner, text="Metal must be spelled exactly as AU, AG, PT or PD",
            font=('TkDefaultFont', 9), bg="#fffbf0", fg="#555555",
            anchor='w').pack(fill='x')
        Label(warn_inner, text="For multiple metals, submit and repeat process",
            font=('TkDefaultFont', 9), bg="#fffbf0", fg="#555555",
            anchor='w').pack(fill='x')

        self.client = self.combobox.get()

        def submit():
            client = 'New client'
            metal = metal_var.get()
            discount = float(discount_var.get())
            if client in self.client_rates:
                self.client_rates[client][metal] = discount
            else:
                self.client_rates[client] = {metal: discount}
            top.destroy()

        submit_button = Button(frame, text="Submit information", command=submit,
                            font=('TkDefaultFont', 10), relief='flat',
                            bg="#e8e8e8", activebackground="#d4d4d4",
                            cursor="hand2", pady=6)
        submit_button.grid(column=0, row=5, columnspan=2, sticky='ew', pady=(14, 14))
        

App()
