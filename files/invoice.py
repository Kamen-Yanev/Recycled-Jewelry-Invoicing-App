import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date
import os
import tempfile
import subprocess
import shutil
import platform
from tkinter import filedialog

from drive import upload_to_drive
from vat import populate_vat
from utils import (
    metal_colors_dict,
    MAIN_TABLE_ROWS,
    INVOICE_MARKET_RATES,
    INVOICE_DISCOUNT_FIELDS,
    CLIENT_INFORMATION_FIELDS,
)
from config import BALANCE_FILE, INVOICE_TEMPLATE


def customers():
    """Returns a dict of client name -> {metal: discount rate}. 'Rate of day' == 0."""
    df = pd.read_excel(BALANCE_FILE, sheet_name='Sheet3', dtype=str)
    df = df.apply(lambda x: x.str.strip() if x.dtype == 'object' or x.dtype == 'string' else x)
    df.columns = ['Client', 'AU', 'AG', 'PT', 'PD']

    clients = {}
    for client in range(len(df)):
        clients[f'{df.iloc[client,0]}'] = {}
        for metal in range(1, 5):
            if df.iloc[client, metal] == 'Rate of day':
                clients[f'{df.iloc[client,0]}'][f'{df.columns[metal]}'] = 0
            elif pd.isna(df.iloc[client, metal]):
                continue
            else:
                clients[f'{df.iloc[client,0]}'][f'{df.columns[metal]}'] = float(df.iloc[client, metal])
    return clients


def isolate_cpm_table(og_df):
    """Extracts the CPM table from the raw dataframe. Returns a list of dicts."""
    necessary_info = []

    locate_start = np.where(og_df.values == 'NO:')
    start_row = int(locate_start[0][0])

    locate_end = np.where(og_df.values == 'SUB TOTAL')
    end_row = int(locate_end[0][0])

    df = og_df[start_row:end_row - 1]
    df = df.iloc[:, 1:]

    df.columns = df.iloc[0]
    df = df.iloc[1:]
    df = df[df["NO:"].notna()]
    df = df.drop(columns=df.columns[[6, 7, 9]])
    df.columns = ['Number', 'Description', 'Melted Weight', 'Metal', 'ASSAY %', 'Fine Metal', 'Rate']
    df = df.astype({'Fine Metal': float, 'Rate': float})

    for i in range(len(df)):
        necessary_info.append(df.iloc[i].to_dict())
    return necessary_info


def isolate_pn(og_df):
    """Extracts the PN number from the raw dataframe. Returns it as a string."""
    pn_position = np.where(og_df.values == 'PN')
    pn_row = int(pn_position[0][0])
    pn_col = int(pn_position[1][0])
    pn = og_df.iloc[pn_row, pn_col + 1]
    return pn


def invoice(root, cpm_doc, client, vat_path, client_rates):
    """
    Main invoice generation function.
    Reads CPM doc, fills invoice template, saves PDF + Excel locally,
    then uploads PDF and VAT form to Google Drive.
    """
    og_df = pd.read_excel(cpm_doc, dtype=str)
    og_df = og_df.apply(lambda x: x.str.strip() if x.dtype == 'string' or x.dtype == 'object' else x)

    cpm_table = isolate_cpm_table(og_df)
    pn = isolate_pn(og_df)

    inv_df = pd.read_excel(INVOICE_TEMPLATE, sheet_name='CPM purchase Invoice', dtype=str)
    inv_df = inv_df.apply(lambda x: x.apply(lambda v: ' '.join(str(v).split()) if isinstance(v, str) else v) if x.dtype == 'object' or x.dtype == 'string' else x)
    starting_row = int(np.where(inv_df.values == 'Melted Weight')[0][0] + 3)

    wb = load_workbook(INVOICE_TEMPLATE, data_only=True)
    ws = wb['CPM purchase Invoice']
    column_names = 'BCDEFG'
    table_cols = starting_row - 1

    ws['H12'] = f'PN{pn}'
    ws['H11'] = date.today()

    for dict in cpm_table:
        for col in column_names:
            cell_name = ws[f'{col}{table_cols}'].value
            corrected_cell_name = ' '.join(str(cell_name).split())
            ws[f'{col}{starting_row}'] = dict[corrected_cell_name]
        starting_row += 1

    cpm_metals = {}
    for dict in cpm_table:
        if dict['Rate'] in cpm_metals:
            continue
        else:
            cpm_metals[dict['Metal']] = float(dict['Rate'])

    for metal in cpm_metals:
        if metal in INVOICE_MARKET_RATES:
            ws[INVOICE_MARKET_RATES[metal]] = cpm_metals[metal]

    for metal in cpm_metals:
        if metal in client_rates[client]:
            if metal in INVOICE_DISCOUNT_FIELDS:
                if client_rates[client][metal] == 0:
                    ws[INVOICE_DISCOUNT_FIELDS[metal]] = 'Rate of the day'
                    ws[CLIENT_INFORMATION_FIELDS[metal]] = f'R{cpm_metals[metal]}'
                else:
                    ws[INVOICE_DISCOUNT_FIELDS[metal]] = f'{client_rates[client][metal]*100:.2f}%'
                    ws[CLIENT_INFORMATION_FIELDS[metal]] = f'R{cpm_metals[metal]*client_rates[client][metal] + cpm_metals[metal]:,.2f}'

    grand_total = 0

    for i, dict in enumerate(cpm_table):
        row = MAIN_TABLE_ROWS[i]
        print(f"i={i}, row={row}, E value={ws[row['metal']].value}, F value={ws[row['assay']].value}")
        client_rate = cpm_metals[dict["Metal"]] * client_rates[client][dict["Metal"]] + cpm_metals[dict["Metal"]]

        ws[row['rate']] = f'R{client_rate:,.2f}'
        ws[row['total']] = f'R{dict["Fine Metal"] * client_rate:,.2f}'
        grand_total += dict['Fine Metal'] * client_rate

        assay_value = round(float(ws[row['assay']].value) * 100, 2)
        ws[row['assay']] = f'{assay_value:.2f}%'

        metal_value = ws[row['metal']].value
        if metal_value in metal_colors_dict:
            ws[row['metal']].fill = metal_colors_dict[metal_value]

    ws['I34'] = f'R{grand_total:,.2f}'

    # Build VAT fields from invoice data
    description = ws['E24'].value
    quantity_mass = ''
    counter = len(cpm_table)
    for dict in cpm_table:
        counter -= 1
        quantity_mass += f'{dict["Description"]}: {dict["Fine Metal"]:,.2f}'
        if counter > 0:
            quantity_mass += ', '

    # Saving the file
    sheets_to_delete = [s for s in wb.sheetnames if s != 'CPM purchase Invoice']
    for sheet in sheets_to_delete:
        del wb[sheet]

    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
    ws.sheet_properties = WorksheetProperties()
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.print_area = 'B1:I36'

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    pdf_path = filedialog.asksaveasfilename(
        parent=root,
        defaultextension='.pdf',
        filetypes=[('PDF files', '*.pdf')],
        title='Save invoice as PDF',
        initialfile=f'PN{pn}'
    )

    xlsx_path = os.path.splitext(pdf_path)[0] + '.xlsx'

    if pdf_path:
        if platform.system() == 'Windows':
            libreoffice_path = r'C:\Program Files\LibreOffice\program\soffice.exe'
        else:
            libreoffice_path = '/Applications/LibreOffice.app/Contents/MacOS/soffice'

        subprocess.run([
            libreoffice_path,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', os.path.dirname(pdf_path),
            tmp_path
        ])
        converted = os.path.join(os.path.dirname(pdf_path), os.path.basename(tmp_path).replace('.xlsx', '.pdf'))
        os.rename(converted, pdf_path)

        try:
            vat_buffer = None
            if vat_path and os.path.exists(vat_path):
                vat_buffer = populate_vat(vat_path, description, quantity_mass, pn, grand_total)
            else:
                print(f"No VAT template found for {client}, skipping VAT form.")
            upload_to_drive(pdf_path, pn, vat_buffer)
        except Exception as e:
            print(f"Drive upload failed: {e}")

    if xlsx_path:
        shutil.copy(tmp_path, xlsx_path)

    os.remove(tmp_path)
