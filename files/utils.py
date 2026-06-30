from openpyxl.styles import PatternFill

metal_colors_dict = {
    'AU': PatternFill(start_color='d3af37', end_color='d3af37', fill_type='solid'),
    'AG': PatternFill(start_color='e6e6e6', end_color='e6e6e6', fill_type='solid'),
    'PD': PatternFill(start_color='989da1', end_color='989da1', fill_type='solid'),
    'PT': PatternFill(start_color='e5e4e2', end_color='e5e4e2', fill_type='solid')
}

# Each dict groups the cells for one row of the main CPM table in the invoice.
# Rows are ordered top-to-bottom (row 24 through row 33).
MAIN_TABLE_ROWS = [
    {'metal': 'E24', 'assay': 'F24', 'rate': 'H24', 'total': 'I24'},
    {'metal': 'E25', 'assay': 'F25', 'rate': 'H25', 'total': 'I25'},
    {'metal': 'E26', 'assay': 'F26', 'rate': 'H26', 'total': 'I26'},
    {'metal': 'E27', 'assay': 'F27', 'rate': 'H27', 'total': 'I27'},
    {'metal': 'E28', 'assay': 'F28', 'rate': 'H28', 'total': 'I28'},
    {'metal': 'E29', 'assay': 'F29', 'rate': 'H29', 'total': 'I29'},
    {'metal': 'E30', 'assay': 'F30', 'rate': 'H30', 'total': 'I30'},
    {'metal': 'E31', 'assay': 'F31', 'rate': 'H31', 'total': 'I31'},
    {'metal': 'E32', 'assay': 'F32', 'rate': 'H32', 'total': 'I32'},
    {'metal': 'E33', 'assay': 'F33', 'rate': 'H33', 'total': 'I33'},
]

INVOICE_MARKET_RATES = {'AU': 'C15', 'AG': 'C16', 'AG2': 'C17', 'PT': 'C18', 'PD': 'C19'}
INVOICE_DISCOUNT_FIELDS = {'AU': 'D15', 'AG': 'D16', 'AG2': 'D17', 'PT': 'D18', 'PD': 'D19'}
CLIENT_INFORMATION_FIELDS = {'AU': 'I17', 'AG': 'I18', 'AG2': 'I19', 'PT': 'I20', 'PD': 'I21'}
